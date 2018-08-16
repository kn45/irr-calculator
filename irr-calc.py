#!/usr/bin/env python
# -*- coding=utf8 -*-

from collections import OrderedDict
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


WB_PTH = './irr_calculator.xlsx'
WB_PTH2 = './irr_calculator_output.xlsx'

wbi = load_workbook(WB_PTH)
wsi = wbi.worksheets[0]

irr_data = {}

srcs = {
    u'投资': 1,
    u'支出': 2,
    u'分红': 3,
    u'估值': 4,
    u'退出': 5}

"""
fill_yellow = PatternFill(fill_type=None,
    start_color='FFFFFF00',
    end_color='FF000000')
"""

fill_yellow = PatternFill('solid', fgColor="FFFF00")

for nrow, row in enumerate(wsi.rows):
    if nrow == 0:
        continue
    if row[0].value is None:
        break
    corp1 = row[0].value
    corp2 = row[1].value

    for src in srcs:
        offset = srcs[src]
        amount = row[2*offset + 1].value
        try:
            x = float(amount)
        except:
            continue
        if amount != 0.0:
            dt = row[2*offset].value
            if corp2 not in irr_data:
                irr_data[corp2] = {}
            if src not in irr_data[corp2]:
                irr_data[corp2][src] = []
            irr_data[corp2][src].append([corp1, amount, dt])

wso = wbi.create_sheet('summary')
wso.append(['']*5 + [u'折现率:', 0.1])
wso.cell(row=wso.max_row, column=6).fill = fill_yellow
wso.cell(row=wso.max_row, column=7).fill = fill_yellow
for corp2 in irr_data:
    nln = 0
    wso.append(['', corp2])
    wso.cell(row=wso.max_row, column=2).fill = fill_yellow
    for src, details in sorted(irr_data[corp2].items(), key=lambda x:srcs[x[0]]):
        for detail in details:
            wso.append([src] + detail)
            wso.cell(row=wso.max_row, column=3).number_format = '#,##0'
            wso.cell(row=wso.max_row, column=4).number_format = 'yyyy-mm-dd'
            nln += 1
        wso.merge_cells(
            start_row=wso.max_row-len(details)+1, start_column=1, end_row=wso.max_row, end_column=1)
    ed_row = wso.max_row
    st_row = ed_row - nln + 1
    formula = '=XIRR(C%d:C%d, D%d:D%d, G1)' % (st_row, ed_row, st_row, ed_row)
    wso.append(['IRR', formula])
    wso.cell(row=wso.max_row, column=2).number_format = '0.00%'
    wso.cell(row=wso.max_row, column=1).fill = fill_yellow
    wso.cell(row=wso.max_row, column=2).fill = fill_yellow
    for _ in range(3):
        wso.append([''])


wbi.save(WB_PTH2)

"""
corp2_1:
    invest: [['c1_1', amount, dt], ['c1_2', amount, dt]]
    fenhong:
    tuichu:

corp2_2:
    invest: []
    fenhong: []
    tuichu: []
"""
